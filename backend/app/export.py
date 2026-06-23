from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


POLLUTANTS = ["CO", "THC", "NOx", "CO2", "CH4", "NMHC", "PM", "PN"]


def export_xlsx(tests: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Emission Compilation"
    headers = [
        "Status", "Test Date", "Program", "Cycle", "Configuration", "Transmission", "Lab",
        "Vehicle", "VN No.", "Catalyst", "STT", "Start SOC (%)", "ODO (km)", "Inertia (kg)",
        *[
            header
            for pollutant in POLLUTANTS
            for header in ([f"{pollutant} (#/km)"] if pollutant == "PN" else [f"{pollutant} (mg/km)", f"{pollutant} (g/km)"])
        ],
        "Source Result Unit", "PDF Source", "XLSM Source", "Review Flags",
    ]
    ws.append(headers)
    for test in tests:
        results = test.get("results", {})
        source = test.get("source", {})
        emission_values = []
        for pollutant in POLLUTANTS:
            value = results.get(pollutant)
            if pollutant == "PN":
                emission_values.append(value)
            else:
                emission_values.extend([value, value / 1000 if value is not None else None])
        ws.append([
            test.get("status"), test.get("date"), test.get("project"), test.get("cycle"), test.get("config"),
            test.get("transmission"), test.get("lab"), test.get("vehicleModel"), test.get("vnNo"),
            test.get("catalystState"), test.get("stt"), test.get("startSoc"), test.get("odo"),
            test.get("inertia"), *emission_values, (test.get("units") or {}).get("resultsSource", "mg/km"), source.get("pdf"),
            source.get("xlsm"), ", ".join(test.get("lowConfidence", [])),
        ])
    fill = PatternFill("solid", fgColor="12313A")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, header in enumerate(headers, 1):
        width = 16
        if header in {"Vehicle", "Catalyst", "PDF Source", "XLSM Source", "Review Flags"}:
            width = 30
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 34
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
